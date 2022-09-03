import requests
import openpyxl
from bs4 import BeautifulSoup

#Storing Data in Excel
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Top Rated Movies"
print(excel.sheetnames)
sheet.append(["Ranking", "Movie Title", "IMDB Rating", "Release Year"])


#Object to get URL
req= requests.get("https://www.imdb.com/chart/top/")
#Processing url contents with html.parser
cont = req.content
soup = BeautifulSoup(cont, "html.parser")

#Finding the class that all movies fall under in HTML inspect
all = soup.find("tbody", class_="lister-list").find_all("tr")

#Creating For Loop to iterate through each listing 
for movie in all:
    #Creating objects for the details of the movies
    rank = movie.find("td", class_="titleColumn").get_text(strip=True).split(".")[0]
    title = movie.find("td", class_="titleColumn").a.text #.a only gives the first peice of information
    rating_a = movie.find("td", class_="ratingColumn imdbRating").strong.text
    rating = rating_a + "/10"
    year = movie.find("td", class_="titleColumn").span.text.strip("()")
    
    movies = [rank, title, rating, year]
    print(movies)
    sheet.append(movies)

excel.save("Top Rated Movies (IMDB).xlsx")  
    
